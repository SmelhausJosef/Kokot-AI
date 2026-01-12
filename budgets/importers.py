from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from typing import Dict, Iterable, Optional

from openpyxl import load_workbook

from .models import Budget, BudgetHeader, BudgetItem


class ExcelImportError(Exception):
    pass


HEADER_TYPES = {
    "Stavba": 1,
    "Skupina objektů": 2,
    "Objekt": 3,
    "Podobjekt": 4,
    "Oddíl": 5,
}


@dataclass(frozen=True)
class ColumnMap:
    type_idx: int
    code_idx: Optional[int]
    description_idx: int
    unit_idx: Optional[int]
    unit_price_idx: Optional[int]


def import_budget_from_excel(budget: Budget) -> int:
    if not budget.excel_file:
        raise ExcelImportError("Budget has no Excel file to import.")

    workbook = load_workbook(budget.excel_file.path, data_only=True)
    if "Zakázka" not in workbook.sheetnames:
        raise ExcelImportError("Excel sheet 'Zakázka' not found.")

    sheet = workbook["Zakázka"]
    header_row, column_map = find_header_row(sheet.iter_rows(values_only=True))

    header_stack: Dict[int, BudgetHeader] = {}
    created_items = 0
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        row_type = get_cell_text(row, column_map.type_idx)
        if not row_type:
            if is_measurement_line(row, column_map):
                continue
            continue

        if row_type in HEADER_TYPES:
            title = get_cell_text(row, column_map.description_idx)
            if not title:
                continue
            level = HEADER_TYPES[row_type]
            parent = None
            for candidate_level in sorted(header_stack.keys(), reverse=True):
                if candidate_level < level:
                    parent = header_stack[candidate_level]
                    break
            header = BudgetHeader.objects.create(budget=budget, parent=parent, title=title)
            header_stack[level] = header
            for deeper_level in [key for key in header_stack.keys() if key > level]:
                header_stack.pop(deeper_level, None)
            continue

        if row_type == "SUB":
            if not header_stack:
                raise ExcelImportError("SUB row encountered before any header row.")
            header = header_stack[max(header_stack.keys())]
            code = get_cell_text(row, column_map.code_idx)
            description = get_cell_text(row, column_map.description_idx)
            if not description:
                continue
            measure_unit = get_cell_text(row, column_map.unit_idx)
            unit_price_value = None
            if column_map.unit_price_idx is not None and column_map.unit_price_idx < len(row):
                unit_price_value = row[column_map.unit_price_idx]
            unit_price = parse_decimal(unit_price_value)
            BudgetItem.objects.create(
                header=header,
                code=code,
                description=description,
                measure_unit=measure_unit,
                price_for_unit=unit_price,
            )
            created_items += 1

    return created_items


def find_header_row(rows: Iterable[Iterable[object]]) -> tuple[int, ColumnMap]:
    for index, row in enumerate(rows, start=1):
        header_map = normalize_headers(row)
        if "Typ" in header_map and "Popis" in header_map:
            return index, ColumnMap(
                type_idx=header_map["Typ"],
                code_idx=header_map.get("Kód"),
                description_idx=header_map["Popis"],
                unit_idx=header_map.get("MJ"),
                unit_price_idx=header_map.get("Jedn. Cena"),
            )
    raise ExcelImportError("Header row with 'Typ' and 'Popis' not found.")


def normalize_headers(row: Iterable[object]) -> Dict[str, int]:
    header_map: Dict[str, int] = {}
    for idx, value in enumerate(row):
        if value is None:
            continue
        name = str(value).strip()
        if not name:
            continue
        header_map[name] = idx
    return header_map


def get_cell_text(row: Iterable[object], idx: Optional[int]) -> str:
    if idx is None:
        return ""
    value = row[idx] if idx < len(row) else None
    if value is None:
        return ""
    return str(value).strip()


def parse_decimal(value: object) -> Decimal:
    if value is None or value == "":
        return Decimal("0.00")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = str(value).strip()
    if not text:
        return Decimal("0.00")
    normalized = text.replace("\xa0", "").replace(" ", "").replace(",", ".")
    try:
        return Decimal(normalized)
    except InvalidOperation as exc:
        raise ExcelImportError(f"Invalid decimal value: {value}") from exc


def is_measurement_line(row: Iterable[object], column_map: ColumnMap) -> bool:
    for idx in [column_map.code_idx, column_map.description_idx]:
        text = get_cell_text(row, idx)
        if "Výkaz výměr:" in text or "Ztratné:" in text:
            return True
    return False
